#!/usr/bin/env python3
"""
Generates the SONIS event-import workbook (tmi_events format) plus the matching
admin-portal event payload, from a simple semester event list.

Input CSV columns (header required, order free):
    name       - event name, e.g. "LLSA: The Importance of Legal Interpreters"
    date       - event date (MM/DD/YYYY or YYYY-MM-DD)
    type       - PS, CC, or BOTH (dual-purpose: gets a PS row AND a CC row)

Usage:
    python3 generate_sonis_events.py events.csv --next-id 528 \
        --xlsx sonis_import.xlsx --portal portal_events.json

SONIS assigns row ids on import (+1 per row, no id column in the template), so
ROW ORDER in the workbook is what makes the predicted ids real. Rows are
emitted in the exact order ids are assigned; dual events emit the PS row first,
then the CC row, matching existing data (e.g. LLSA = PS 511 + CC 512).
After importing, spot-check in SONIS that the first new event landed on the
expected id before exporting any attendance.
"""
import argparse, csv, json, sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

HEADERS = ["EV_SDESC", "EV_LDESC", "EV_ST_DATE", "EV_SP_DATE",
           "EV_DISABLE", "EV_TARGET", "RESTRICTED", "INTPL_BIT"]

def parse_date(raw):
    raw = raw.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unparseable date: {raw!r}")

import re

def sanitize(text):
    text = text.replace("&", " and ")
    text = re.sub(r"[^A-Za-z0-9 ]", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def short_desc(prefix, date):
    # Matches existing SONIS convention: "PS 2026 04/16" (nchar 35).
    return f"{prefix} {date.year} {date.strftime('%m/%d')}"

def long_desc(prefix, name):
    text = f"{prefix}: {name}"
    if len(text) > 150:
        text = text[:150]  # ev_ldesc is nchar(150)
    return text

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("csv_path")
    ap.add_argument("--next-id", type=int, required=True,
                    help="Next free SONIS event row id (last existing id + 1)")
    ap.add_argument("--xlsx", default="sonis_event_import.xlsx")
    ap.add_argument("--portal", default="portal_events.json")
    ap.add_argument("--strip-special", action="store_true",
                    help="SONIS's import validator rejects punctuation in the "
                         "nchar text columns (slash, colon, parens, ampersand) "
                         "even though the UI accepts them. Strips text fields "
                         "to alphanumerics and spaces; restore punctuation in "
                         "the SONIS UI after importing.")
    args = ap.parse_args()

    rows = list(csv.DictReader(open(args.csv_path, newline="", encoding="utf-8-sig")))
    norm = lambda r, k: (r.get(k) or r.get(k.title()) or r.get(k.upper()) or "").strip()

    events = []
    for r in rows:
        name, date_raw, kind = norm(r, "name"), norm(r, "date"), norm(r, "type").upper()
        if not name or not date_raw:
            continue
        if kind not in ("PS", "CC", "BOTH"):
            sys.exit(f"Row {name!r}: type must be PS, CC, or BOTH (got {kind!r})")
        events.append({"name": name, "date": parse_date(date_raw), "kind": kind})

    # Chronological, so numbering follows the calendar.
    events.sort(key=lambda e: e["date"])

    next_id = args.next_id
    sonis_rows, portal_events = [], []
    for ev in events:
        prefixes = ["PS", "CC"] if ev["kind"] == "BOTH" else [ev["kind"]]
        ids = {}
        for prefix in prefixes:  # PS row first for dual events
            ids[prefix] = next_id
            # Every value as plain text. The SONIS schema declares the date
            # columns as datetime with Max Length 8, and its import validator
            # applies that as a string-length cap — so dates must be the
            # 8-character MM/DD/YY form. Typed datetime cells fail outright
            # (stringified with a time component: wrong length, plus ':' and
            # '-' tripping the special-character check).
            date_text = ev["date"].strftime("%m/%d/%y")
            sdesc = short_desc(prefix, ev["date"])
            ldesc = long_desc(prefix, ev["name"])
            if args.strip_special:
                sdesc, ldesc = sanitize(sdesc), sanitize(ldesc)
            sonis_rows.append([
                sdesc, ldesc,
                date_text, date_text,
                "0", "0.00", "0", "0",
            ])
            next_id += 1

        primary = ids.get("PS") or ids.get("CC")
        portal_events.append({
            "eventNumber": primary,
            "name": ev["name"],
            "date": ev["date"].strftime("%Y-%m-%d"),
            "isActive": True,
            "isCCC": ev["kind"] == "BOTH",
            "cccEventId": ids.get("CC") if ev["kind"] == "BOTH" else None,
            "description": f"{'/'.join(prefixes)} event",
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    font = Font(name="Arial", size=10)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(name="Arial", size=10, bold=True)
    for row in sonis_rows:
        ws.append(row)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.font = font
            c.number_format = "@"  # text
    for col, width in zip("ABCDEFGH", (16, 70, 12, 12, 10, 10, 10, 10)):
        ws.column_dimensions[col].width = width
    wb.save(args.xlsx)

    json.dump(portal_events, open(args.portal, "w"), indent=2)

    print(f"{len(events)} events -> {len(sonis_rows)} SONIS rows "
          f"(ids {args.next_id}..{next_id - 1})")
    print(f"workbook: {args.xlsx}")
    print(f"portal payload: {args.portal}")
    for ev, pe in zip(events, portal_events):
        tag = f"PS {pe['eventNumber']}" + (f" + CC {pe['cccEventId']}" if pe["isCCC"] else "")
        print(f"  {ev['date'].strftime('%m/%d/%Y')}  [{tag:>16}]  {ev['name'][:60]}")

if __name__ == "__main__":
    main()
